import os,json, time, openpyxl
from datetime import datetime
from selenium import webdriver
from sys import platform
import pathlib
from pathlib import Path
from openpyxl import Workbook
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait,Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import random
from random import randint
from selenium.webdriver.common.action_chains import ActionChains



import luu_auto_function

chrome_options = webdriver.ChromeOptions()

class objects:
    now = datetime.now()
    year = now.strftime("%Y")
    month = now.strftime("%m")
    day = now.strftime("%d")
    time1 = now.strftime("%H:%M:%S")
    date_time = now.strftime("%Y/%m/%d, %H:%M:%S")
    date_id = date_time.replace("/", "").replace(", ", "").replace(":", "")[2:]
    testcase_pass = "Test case status: pass"
    testcase_fail = "Test case status: fail"

if platform == "linux" or platform == "linux2":
    local = "/home/oem/groupware-auto-test"
    json_file = local + "/luu_settingluu.json"
    with open(json_file) as json_data_file:
        data = json.load(json_data_file)
    driver = webdriver.Chrome("/usr/bin/chromedriver")
    file_buildr = local+"/Attachment/form_data_builder.xls"
    file_img = local+"/Attachment/signature_mage.jpg"
    file_editor = local+"/Attachment/upload_office file_editor.xls"
    local = "/home/oem/groupware-auto-test"
    log_folder = "/Log/"
    execution_log = local + log_folder + "execution_log_" + str(objects.date_id) + ".txt"
    #fail_log = execution_log.replace("execution_log_", "fail_log_")
    #error_log = execution_log.replace("execution_log_", "error_log_") 
else :
    local = "D:\File_Du_Lieu\Automation Test\luungo_automationtest"
    json_file = local + "\\luu_settingluu.json"
    with open(json_file) as json_data_file:
        data = json.load(json_data_file)
    driver = webdriver.Chrome(local + "\\chromedriver.exe")
    log_folder = "\\Log\\"
    execution_log = local + log_folder + "execution_log_" + str(objects.date_id) + ".txt"
    #fail_log = execution_log.replace("execution_log_", "fail_log_")
    #error_log = execution_log.replace("execution_log_", "error_log_")
    file_buildr=local+"\\attachment\\form_data_builder.xls"
    file_img= local+"\\attachment\\signature_mage.jpg"
    file_editor=local+"\\attachment\\upload_office file_editor.xls"

testcase_log = local + log_folder + "testcase_luu_ngo_gw_" + str(objects.date_id) + ".xlsx"   
logs = [execution_log,testcase_log]
#logs = [execution_log,fail_log,error_log,testcase_log]
for log in logs: 
    if".txt" in log:
        open(log,"x").close()
    else:
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1).value = "Menu"
        ws.cell(row=1, column=2).value = "Sub-Menu"
        ws.cell(row=1, column=3).value = "Test Case Name"
        ws.cell(row=1, column=4).value = "Status"
        ws.cell(row=1, column=5).value = "Description"
        ws.cell(row=1, column=6).value = "Date"
        ws.cell(row=1, column=7).value = "Tester"
        wb.save(log)




'''
# create log file of fail test case
open(execution_log, "x").close()

# create log file of fail test case
open(fail_log, "x").close()

# create log file of fail test case
open(error_log, "x").close()

open(testcase_log, "x").close()
'''

def Logging(msg):
    print(msg)
    log_msg = open(execution_log, "a")
    log_msg.write(str(msg) + "\n")
    log_msg.close()

def ValidateFailResultAndSystem(fail_msg):
    print(fail_msg)
    #append_fail_result = open(fail_log, "a")
    append_fail_result = open("a")
    append_fail_result.write("[FAILED TEST CASE] " + str(fail_msg) + "\n")
    append_fail_result.close()


def TestCase_LogResult(menu, sub_menu, testcase, status, description, tester):
    Logging("description")
    if status=="Pass":
        print(objects.testcase_pass)
    else:
        print(objects.testcase_fail)

    wb = openpyxl.load_workbook(testcase_log)
    current_sheet = wb.active
    start_row = len(list(current_sheet.rows))+1
    current_sheet.cell(row=start_row,column=1).value=menu
    current_sheet.cell(row=start_row,column=2).value=sub_menu
    current_sheet.cell(row=start_row,column=3).value=testcase
    current_sheet.cell(row=start_row,column=4).value=status
    current_sheet.cell(row=start_row,column=5).value=description
    current_sheet.cell(row=start_row,column=6).value=objects.date_time
    current_sheet.cell(row=start_row,column=7).value= tester
    
    wb.save(testcase_log)




# start login page
driver.implicitly_wait(5)
driver.set_window_size(1024, 600)
driver.maximize_window()


def access_qa(domain_name):
    driver.get(domain_name)
    print(domain_name)
    print("- Access login page")
    driver.find_element_by_id("log-userid").send_keys(data["user_name"])
    print("- Input user ID")
    frame_element = driver.find_element_by_id("iframeLoginPassword")
    driver.switch_to.frame(frame_element)
    driver.find_element_by_id("p").send_keys(data["user_password"])
    driver.switch_to.default_content()
    print("- Input user password")
    driver.find_element_by_id("btn-log").send_keys(Keys.RETURN)
    print("- Click button Login")

def close_pop_up():
    try:
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["turn_off"]))).send_keys(Keys.RETURN)
        print("- Close pop up Successfully")
    except:
        print("- Don't show pop up")


class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'

def Yellow(msg):
    string_output = bcolors.WARNING + str(msg) + bcolors.ENDC
    return string_output


def Green(msg):
    string_output = bcolors.OKGREEN + str(msg) + bcolors.ENDC
    return string_output
    
def Red(msg):
    string_output = bcolors.FAIL + str(msg) + bcolors.ENDC
    return string_output 
    


class Commands():
    def WaitElementLoaded(time, xpath):
        WebDriverWait(driver, time).until(EC.presence_of_element_located((By.XPATH, xpath)))

    def Wait10s_ClickElement(xpath):
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, xpath)))
        element = driver.find_element_by_xpath(xpath)
        element.click()
        return element

    def Wait10s_InputElement(xpath, value):
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, xpath)))
        element = driver.find_element_by_xpath(xpath)
        element.send_keys(value)

    def scroll_view(xpath):
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, xpath)))
        element = driver.find_element_by_xpath(xpath)
        element.location_once_scrolled_into_view
        return element


    def drag_drop_Element(xpath, xpath2):
        '''• Usage: Move to view element by ActionChains
                return WebElement'''

        element = driver.find_element_by_xpath(xpath)
        element_1 = driver.find_element_by_xpath(xpath2)
        actions = ActionChains(driver)
        #actions.move_to_element(element)
        actions.drag_and_drop(element,element_1)
        actions.perform()
        time.sleep(1)

        return element
    
    def Wait10s_InputElement_return(xpath, value):
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, xpath)))
        element = driver.find_element_by_xpath(xpath)
        element.send_keys(value)
        element.send_keys(Keys.RETURN)

    
    def FindElement(xpath):
        element = driver.find_element_by_xpath(xpath)

        return element

    def FindElements(xpath):
        element = driver.find_elements_by_xpath(xpath)

        return element

    def SwitchToFrame(frame_xpath):
        WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.XPATH, frame_xpath)))
        frame = Commands.FindElement(frame_xpath)
        driver.switch_to.frame(frame)
        return frame

    def SwitchToFrame_no(frame_xpath):
        WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CLASS_NAME, frame_xpath)))
        frame = driver.find_element_by_class_name(frame_xpath)
        driver.switch_to.frame(frame)
    
    def Selectbox_ByVisibleText(xpath, selected_text):
        
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, xpath)))
        element = Select(driver.find_element_by_xpath(xpath))
        element.select_by_visible_text(selected_text)

    def Wait10s_clearElement(xpath):
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, xpath)))
        element = driver.find_element_by_xpath(xpath)
        element.clear()
    
    def Wait10s_hold_moveElement(xpath,xpath2):
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, xpath)))
        element = driver.find_element_by_xpath(xpath)
        element_1 = driver.find_element_by_xpath(xpath2)
        actions = ActionChains(driver)
        actions.click_and_hold(element).move_to_element(element_1)
        actions.perform()

    def GetListlength(xpath):
        
        element = int(len(driver.find_elements_by_xpath(xpath)))


    def RemoveDuplicate_fromList(selected_list):
    
        #selected_list = list(dict.fromkeys(selected_list))
        ActionChains(driver).key_down(Keys.CONTROL).send_keys('selected_list').key_up(Keys.CONTROL).perform()